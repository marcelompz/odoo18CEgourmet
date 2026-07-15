# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
import base64
import io
from datetime import datetime, timedelta

class AttendanceReportWizard(models.TransientModel):
    _name = 'attendance.report.wizard'
    _description = 'Attendance Report Wizard'
    
    report_type = fields.Selection([
        ('excel', 'Excel Report'),
        ('pdf', 'PDF Report')
    ], string='Report Type', default='excel', required=True)
    start_date = fields.Date(string='Start Date', required=True, default=lambda self: fields.Date.today() - timedelta(days=30))
    end_date = fields.Date(string='End Date', required=True, default=lambda self: fields.Date.today())
    device_id = fields.Many2one('biometric.device', string='Device')
    employee_ids = fields.Many2many('hr.employee', string='Employees')
    department_id = fields.Many2one('hr.department', string='Department')
    group_by_employee = fields.Boolean(string='Group by Employee', default=True)
    include_processed = fields.Boolean(string='Include Processed Logs', default=True)
    include_unprocessed = fields.Boolean(string='Include Unprocessed Logs', default=True)
    
    @api.constrains('start_date', 'end_date')
    def _check_dates(self):
        for wizard in self:
            if wizard.start_date > wizard.end_date:
                raise ValidationError(_('Start date cannot be after end date'))
    
    def _get_domain(self):
        domain = [
            ('check_time', '>=', self.start_date),
            ('check_time', '<=', self.end_date + ' 23:59:59')
        ]
        
        if self.device_id:
            domain.append(('device_id', '=', self.device_id.id))
        
        if self.employee_ids:
            domain.append(('employee_id', 'in', self.employee_ids.ids))
        
        if self.department_id:
            domain.append(('department_id', '=', self.department_id.id))
        
        if not self.include_processed and not self.include_unprocessed:
            domain.append(('id', '=', False))  # Return no records
        elif not self.include_processed:
            domain.append(('processed', '=', False))
        elif not self.include_unprocessed:
            domain.append(('processed', '=', True))
        
        return domain
    
    def action_generate_report(self):
        self.ensure_one()
        
        logs = self.env['attendance.log'].search(self._get_domain())
        
        if not logs:
            raise ValidationError(_('No attendance logs found for the selected criteria'))
        
        if self.report_type == 'excel':
            return self._generate_excel_report(logs)
        else:
            return self._generate_pdf_report(logs)
    
    def _generate_excel_report(self, logs):
        """Generate Excel report using openpyxl"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Attendance Report'
            
            # Header
            headers = ['Date', 'Time', 'Employee', 'Biometric ID', 'Department', 'Device', 'Check Type', 'Processed']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            row = 2
            for log in logs:
                check_datetime = fields.Datetime.from_string(log.check_time)
                ws.cell(row=row, column=1, value=check_datetime.date())
                ws.cell(row=row, column=2, value=check_datetime.time())
                ws.cell(row=row, column=3, value=log.employee_id.name or '')
                ws.cell(row=row, column=4, value=log.biometric_id or '')
                ws.cell(row=row, column=5, value=log.department_id.name or '')
                ws.cell(row=row, column=6, value=log.device_id.name or '')
                ws.cell(row=row, column=7, value=log.check_type or '')
                ws.cell(row=row, column=8, value='Yes' if log.processed else 'No')
                row += 1
            
            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Create attachment
            filename = f'attendance_report_{self.start_date}_{self.end_date}.xlsx'
            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'datas': base64.b64encode(buffer.read()),
                'res_model': 'attendance.report.wizard',
                'res_id': self.id,
            })
            
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'self',
            }
            
        except ImportError:
            raise ValidationError(_('openpyxl library is required for Excel reports. Please install it.'))
    
    def _generate_pdf_report(self, logs):
        """Generate PDF report (placeholder implementation)"""
        # For now, use Excel as fallback since PDF requires more complex template
        return self._generate_excel_report(logs)