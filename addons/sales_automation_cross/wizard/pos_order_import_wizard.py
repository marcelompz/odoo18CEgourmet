from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import csv
import io
import logging
from datetime import datetime
from collections import defaultdict

_logger = logging.getLogger(__name__)

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

class PosImportLine(models.TransientModel):
    _name = 'pos.import.line'
    _description = 'Línea de Mapeo de Importación POS'

    wizard_id = fields.Many2one('pos.import.wizard', string='Asistente', ondelete='cascade')
    column_name = fields.Char(string='Columna', readonly=True)
    odoo_field = fields.Selection([
        ('ignore', 'Ignorar'),
        ('order_ref', 'Order / Referencia (Agrupador)'),
        ('date', 'Order / Fecha'),
        ('partner', 'Order / Cliente'),
        ('line_product', 'Line / Producto (Nombre/Ref)'),
        ('line_qty', 'Line / Cantidad'),
        ('line_price', 'Line / Precio Unitario'),
        ('payment_method', 'Payment / Método de Pago'),
        ('payment_amount', 'Payment / Importe'),
    ], string='Campo en Odoo', default='ignore', required=True)

class PosImportWizard(models.TransientModel):
    _name = 'pos.import.wizard'
    _description = 'Importar Histórico de Ventas PDV'

    file = fields.Binary(string='Archivo', required=True)
    file_name = fields.Char(string='Nombre del Archivo')
    session_id = fields.Many2one('pos.session', string='Sesión de POS', required=True,
                                domain=[('state', '=', 'opened')])
    delimiter = fields.Selection([
        (',', 'Coma (,)'),
        (';', 'Punto y coma (;)')
    ], string='Delimitador CSV', default=',')
    
    is_excel = fields.Boolean(compute='_compute_is_excel')
    
    line_ids = fields.One2many('pos.import.line', 'wizard_id', string='Mapeo de Columnas')

    @api.depends('file_name')
    def _compute_is_excel(self):
        for rec in self:
            rec.is_excel = rec.file_name and (rec.file_name.endswith('.xls') or rec.file_name.endswith('.xlsx'))

    def action_load_columns(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_("Primero suba un archivo."))
        
        self.line_ids.unlink()
        
        data = base64.b64decode(self.file)
        file_ext = self.file_name.split('.')[-1].lower() if self.file_name else ''
        
        headers = []
        if file_ext in ['xls', 'xlsx']:
            headers = self._get_excel_headers(data, file_ext)
        else:
            headers = self._get_csv_headers(data)
            
        mapping_lines = []
        for header in headers:
            if not header: continue
            odoo_field = 'ignore'
            h_lower = str(header).lower().strip()
            # Sugerencias automáticas inteligentes
            if h_lower in ['fecha', 'date', 'momento']: odoo_field = 'date'
            elif h_lower in ['cliente', 'partner', 'customer']: odoo_field = 'partner'
            elif h_lower in ['producto', 'product', 'articulo', 'artículo', 'artículo del combo']: odoo_field = 'line_product'
            elif h_lower in ['cantidad', 'qty', 'count', 'cant', 'unidades', 'cantidad de entrega']: odoo_field = 'line_qty'
            elif h_lower in ['precio', 'price', 'monto', 'total', 'unitario', 'precio unitario']: odoo_field = 'line_price'
            elif h_lower in ['metodo', 'método', 'método de pago', 'payment method']: odoo_field = 'payment_method'
            elif h_lower in ['importe', 'amount', 'pago']: odoo_field = 'payment_amount'
            elif h_lower in ['referencia', 'ref', 'order ref', 'ref. de la orden']: odoo_field = 'order_ref'
            
            mapping_lines.append((0, 0, {
                'column_name': str(header),
                'odoo_field': odoo_field,
            }))
            
        self.write({'line_ids': mapping_lines})
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'pos.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _get_csv_headers(self, data):
        try:
            content = data.decode('utf-8')
        except UnicodeDecodeError:
            try:
                content = data.decode('latin1')
            except:
                raise UserError(_("No se pudo decodificar el archivo CSV."))
        f = io.StringIO(content)
        reader = csv.reader(f, delimiter=self.delimiter)
        try:
            return next(reader)
        except StopIteration:
            return []

    def _get_excel_headers(self, data, file_ext):
        if file_ext == 'xlsx':
            if not openpyxl: raise UserError(_("Instale openpyxl para archivos .xlsx"))
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
            sheet = wb.active
            return [cell.value for cell in sheet[1] if cell.value]
        else:
            if not xlrd: raise UserError(_("Instale xlrd para archivos .xls"))
            book = xlrd.open_workbook(file_contents=data)
            sheet = book.sheet_by_index(0)
            return [sheet.cell_value(0, i) for i in range(sheet.ncols) if sheet.cell_value(0, i)]

    def action_import(self):
        self.ensure_one()
        if not self.line_ids:
            raise UserError(_("Debe cargar las columnas primero."))

        # Crear mapa inverso: campo_odoo -> columna_archivo
        mapping = {line.odoo_field: line.column_name for line in self.line_ids if line.odoo_field != 'ignore'}
        if not mapping.get('line_product'):
            raise UserError(_("Debe mapear al menos el campo de Producto (Line / Producto)."))

        file_ext = self.file_name.split('.')[-1].lower() if self.file_name else ''
        data = base64.b64decode(self.file)
        
        # Obtener todas las filas
        raw_rows = []
        if file_ext in ['xls', 'xlsx']:
            raw_rows = self._parse_excel_rows(data, file_ext)
        else:
            raw_rows = self._parse_csv_rows(data)

        if not raw_rows:
            raise UserError(_("No hay datos para importar."))

        # Agrupación por Order Reference
        orders_dict = defaultdict(lambda: {'order_vals': {}, 'lines': [], 'payments': []})
        
        for row in raw_rows:
            # Obtener identificador de orden (si no hay ref, generamos uno por fila para que sean órdenes individuales si se desea)
            ref = row.get(mapping.get('order_ref')) or 'NO_REF_%s' % len(orders_dict)
            if mapping.get('order_ref') and not row.get(mapping.get('order_ref')):
                 # Si el usuario mapeó un campo de ref pero está vacío en esta fila, usamos la anterior o generamos nueva?
                 # Odoo nativo suele considerar filas vacías como continuación de la anterior.
                 pass

            order_data = orders_dict[ref]
            
            # Datos de Cabecera (solo la primera vez que vemos la referencia)
            if not order_data['order_vals']:
                order_data['order_vals'] = {
                    'date': row.get(mapping.get('date')),
                    'partner': row.get(mapping.get('partner')),
                }

            # Datos de Línea
            if row.get(mapping.get('line_product')):
                order_data['lines'].append({
                    'product': row.get(mapping.get('line_product')),
                    'qty': float(row.get(mapping.get('line_qty')) or 1),
                    'price': float(row.get(mapping.get('line_price')) or 0),
                })

            # Datos de Pago
            if row.get(mapping.get('payment_method')):
                order_data['payments'].append({
                    'method': row.get(mapping.get('payment_method')),
                    'amount': float(row.get(mapping.get('payment_amount')) or 0),
                })

        return self._create_grouped_orders(orders_dict)

    def _parse_csv_rows(self, data):
        try:
            content = data.decode('utf-8')
        except:
            content = data.decode('latin1')
        reader = csv.DictReader(io.StringIO(content), delimiter=self.delimiter)
        return [row for row in reader]

    def _parse_excel_rows(self, data, file_ext):
        if file_ext == 'xlsx':
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            return [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]
        else:
            book = xlrd.open_workbook(file_contents=data)
            sheet = book.sheet_by_index(0)
            headers = [sheet.cell_value(0, i) for i in range(sheet.ncols)]
            rows = []
            for r in range(1, sheet.nrows):
                rows.append({headers[c]: sheet.cell_value(r, c) for c in range(sheet.ncols)})
            return rows

    def _create_grouped_orders(self, orders_dict):
        PosOrder = self.env['pos.order']
        Product = self.env['product.product']
        Partner = self.env['res.partner']
        PaymentMethod = self.env['pos.payment.method']
        
        created_count = 0
        for ref, data in orders_dict.items():
            vals = data['order_vals']
            
            # Buscar Cliente
            partner = False
            if vals.get('partner'):
                p_name = str(vals['partner']).strip()
                partner = Partner.search([('name', '=', p_name)], limit=1)
            
            # Fecha
            date_order = fields.Datetime.now()
            if vals.get('date'):
                try:
                    if isinstance(vals['date'], datetime):
                        date_order = vals['date']
                    else:
                        date_order = fields.Datetime.from_string(str(vals['date']))
                except: pass

            # Preparar Líneas
            line_vals = []
            total_amount = 0
            for l in data['lines']:
                p_search = str(l['product']).strip()
                product = Product.search(['|', ('name', '=', p_search), ('default_code', '=', p_search)], limit=1)
                if not product:
                    _logger.warning("Fila ignorada: Producto no encontrado '%s'" % p_search)
                    continue
                
                subtotal = l['qty'] * l['price']
                total_amount += subtotal
                line_vals.append((0, 0, {
                    'product_id': product.id,
                    'qty': l['qty'],
                    'price_unit': l['price'],
                    'price_subtotal': subtotal,
                    'price_subtotal_incl': subtotal,
                }))

            if not line_vals:
                continue

            # Crear Orden
            order = PosOrder.create({
                'session_id': self.session_id.id,
                'partner_id': partner.id if partner else False,
                'date_order': date_order,
                'pos_reference': ref if not ref.startswith('NO_REF_') else False,
                'amount_tax': 0.0,
                'amount_total': total_amount,
                'amount_paid': 0.0, # Se actualizará con pagos
                'amount_return': 0.0,
                'state': 'draft',
                'lines': line_vals,
            })

            # Crear Pagos
            paid_amount = 0
            for p in data['payments']:
                m_search = str(p['method']).strip()
                method = PaymentMethod.search([('name', '=', m_search)], limit=1)
                if not method:
                    _logger.warning("Pago ignorado: Método '%s' no encontrado" % m_search)
                    continue
                
                # Crear pago directo
                self.env['pos.payment'].create({
                    'pos_order_id': order.id,
                    'amount': p['amount'],
                    'payment_method_id': method.id,
                    'payment_date': date_order,
                })
                paid_amount += p['amount']

            # Validar si está pagada para cambiar estado (opcional para histórico)
            if paid_amount >= total_amount:
                order.write({'state': 'paid', 'amount_paid': paid_amount})
            else:
                 order.write({'amount_paid': paid_amount})
                 
            created_count += 1

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Importación Exitosa'),
                'message': _('Se han creado %s órdenes con sus respectivas líneas y pagos.') % created_count,
                'type': 'success',
                'sticky': False,
            }
        }
