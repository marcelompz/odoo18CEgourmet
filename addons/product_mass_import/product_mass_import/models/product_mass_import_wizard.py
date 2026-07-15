# -*- coding: utf-8 -*-

import base64
import io
import unicodedata
from odoo import models, fields, api, _
from odoo.exceptions import UserError

import openpyxl


def normalize_text(text):
    """
    Normaliza texto para comparación: minúsculas, sin tildes, sin espacios extra.
    Ejemplo: "Artículos de Electricidad" → "articulos de electricidad"
    """
    if not text:
        return ''
    # Convertir a minúsculas
    text = text.lower()
    # Eliminar tildes
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    # Eliminar espacios extra y caracteres especiales
    text = ' '.join(text.split())
    return text


def find_best_match_category(category_name, categories_env):
    """
    Busca una categoría existente con nombre similar (fuzzy match).
    Retorna la categoría encontrada o None si no hay coincidencia.
    
    Estrategia:
    1. Búsqueda exacta normalizada (ignora tildes/mayúsculas)
    2. Búsqueda por contención (si el nombre normalizado está contenido)
    """
    if not category_name:
        return None
    
    normalized_input = normalize_text(category_name)
    
    # Buscar todas las categorías y comparar normalizadas
    all_categories = categories_env.search([])
    
    # 1. Búsqueda exacta normalizada
    for categ in all_categories:
        if normalize_text(categ.name) == normalized_input:
            return categ
    
    # 2. Búsqueda por contención (si el input es más largo y contiene el nombre)
    for categ in all_categories:
        normalized_categ = normalize_text(categ.name)
        if normalized_categ and normalized_categ in normalized_input:
            return categ
        if normalized_input and normalized_input in normalized_categ:
            return categ
    
    # 3. Búsqueda con similaridad simple (ratio de caracteres comunes)
    # Usamos una heurística simple: si comparten >80% de palabras
    input_words = set(normalized_input.split())
    for categ in all_categories:
        normalized_categ = normalize_text(categ.name)
        categ_words = set(normalized_categ.split())
        
        if not categ_words or not input_words:
            continue
        
        # Calcular intersección
        common_words = input_words & categ_words
        total_words = input_words | categ_words
        
        if len(total_words) > 0:
            similarity = len(common_words) / len(total_words)
            if similarity >= 0.8:  # 80% de similitud
                return categ
    
    return None


class ProductMassImportWizard(models.TransientModel):
    _name = 'product.mass.import.wizard'
    _description = 'Wizard de Importación Masiva de Productos desde Excel'

    file_data = fields.Binary(string='Archivo Excel (.xlsx)', required=True)
    filename = fields.Char(string='Nombre del Archivo')
    location_id = fields.Many2one(
        'stock.location',
        string='Ubicación de Inventario',
        required=True,
        domain=[('usage', '=', 'internal')],
        help="Ubicación física donde se cargará el stock inicial por defecto."
    )
    state = fields.Selection([
        ('draft', 'Borrador'),
        ('preview', 'Vista Previa'),
        ('done', 'Completado'),
    ], default='draft', string='Estado')
    product_count = fields.Integer(string='Cantidad de Productos', compute='_compute_product_count')
    preview_ids = fields.One2many('product.mass.import.preview', 'wizard_id', string='Vista Previa')

    @api.model
    def default_get(self, fields_list):
        res = super(ProductMassImportWizard, self).default_get(fields_list)
        warehouse = self.env['stock.warehouse'].search([], limit=1)
        if warehouse:
            res['location_id'] = warehouse.lot_stock_id.id
        return res

    def _compute_product_count(self):
        for wizard in self:
            wizard.product_count = len(wizard.preview_ids)

    def action_download_template(self):
        """Download Excel template for product import"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Plantilla Productos'

        # Headers
        headers = [
            'Referencia Interna',
            'Nombre del Producto',
            'Descripción para PdV',
            'Código de Barras',
            'Disponible en PdV',
            'Categoría de Producto',
            'Categoría de PdV',
            'Precio de Venta',
            'Precio de Costo',
            'Cantidad a la Mano',
            'Tipo de Producto',
            'Trazabilidad'
        ]
        ws.append(headers)

        # Example row
        example = [
            'PROD-001',
            'Producto Ejemplo',
            'Descripción para mostrar en punto de venta',
            '7701234567890',
            'VERDADERO',
            'Electrodomésticos',
            'Electrodomésticos',
            100000.00,
            75000.00,
            50,
            'Almacenable',
            'Ninguno'
        ]
        ws.append(example)

        # Column widths
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        file_content = base64.b64encode(output.read())

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/?model=product.mass.import.wizard&id={self.id}&field=file_data&download=true&filename=plantilla_productos.xlsx',
            'target': 'new',
        }

    def action_parse_excel(self):
        """Parse Excel file and show preview - OPTIMIZED with batch barcode validation"""
        self.ensure_one()
        self.preview_ids.unlink()

        data = base64.b64decode(self.file_data)
        wb = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
        sheet = wb.active

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        # OPTIMIZACIÓN: Extraer todos los barcodes del Excel y validar en una sola consulta
        excel_barcodes = []
        for row in rows:
            if row and len(row) > 3 and row[3]:
                barcode = str(row[3]).strip()
                if barcode:
                    excel_barcodes.append(barcode)
        
        # Una sola consulta para todos los barcodes existentes
        existing_barcodes = set()
        if excel_barcodes:
            existing_barcodes = set(self.env['product.product'].search(
                [('barcode', 'in', excel_barcodes)]
            ).mapped('barcode'))
        
        # Procesar filas y detectar duplicados internos
        preview_data = []
        seen_barcodes_in_file = set()
        
        for idx, row in enumerate(rows, start=2):
            if not row or not row[1]:  # row[1] = name
                continue

            error_msgs = []
            default_code = str(row[0]).strip() if row[0] else False
            name = str(row[1]).strip() if row[1] else False
            pos_description = str(row[2]).strip() if row[2] else False
            barcode = str(row[3]).strip() if row[3] else False
            available_in_pos = str(row[4]).upper() in ['VERDADERO', 'TRUE', '1', 'SI'] if row[4] is not None else True
            categ_name = str(row[5]).strip() if row[5] else False
            pos_categ_name = str(row[6]).strip() if row[6] else False
            list_price = float(row[7]) if row[7] else 0.0
            standard_price = float(row[8]) if row[8] else 0.0
            qty_on_hand = float(row[9]) if row[9] else 0.0

            # Product type - ODoo 19: 'consu'=Goods, 'service'=Service, 'combo'=Combo
            product_type = 'consu'  # Default a Bienes (incluye almacenables)
            if row[10]:
                type_val = str(row[10]).lower()
                if type_val in ['servicio', 'service']:
                    product_type = 'service'
                elif type_val in ['combo']:
                    product_type = 'combo'
                # 'almacenable', 'storable', 'product' -> quedan como 'consu' (Goods)

            # Tracking
            tracking = 'none'
            if row[11]:
                track_val = str(row[11]).lower()
                if 'lote' in track_val:
                    tracking = 'lot'
                elif 'serie' in track_val:
                    tracking = 'serial'

            # Validate barcode uniqueness in database
            if barcode:
                if barcode in existing_barcodes:
                    error_msgs.append(f"Código de barras ya existe en Odoo")
                # VALIDACIÓN DE DUPLICADOS INTERNOS
                if barcode in seen_barcodes_in_file:
                    error_msgs.append(f"Código de barras duplicado en este archivo")
                seen_barcodes_in_file.add(barcode)

            # Validate required fields
            if not default_code:
                error_msgs.append("Referencia interna requerida")

            if not name:
                error_msgs.append("Nombre del producto requerido")

            # Validate numeric fields (only if provided)
            if list_price is not None and list_price < 0:
                error_msgs.append("Precio de venta no puede ser negativo")

            if standard_price is not None and standard_price < 0:
                error_msgs.append("Precio de costo no puede ser negativo")

            if qty_on_hand is not None and qty_on_hand < 0:
                error_msgs.append("Cantidad no puede ser negativa")

            error_str = ', '.join(error_msgs) if error_msgs else ''

            preview_data.append((0, 0, {
                'row_number': idx,
                'default_code': default_code or '',
                'name': name or '',
                'pos_description': pos_description or '',
                'barcode': barcode or '',
                'available_in_pos': available_in_pos,
                'categ_name': categ_name or '',
                'pos_categ_name': pos_categ_name or '',
                'list_price': list_price,
                'standard_price': standard_price,
                'qty_on_hand': qty_on_hand,
                'product_type': product_type,
                'tracking': tracking,
                'error_message': error_str,
                'is_valid': len(error_msgs) == 0,
            }))

        self.write({
            'state': 'preview',
            'preview_ids': preview_data,
        })

        valid_count = sum(1 for p in preview_data if p[2]['is_valid'])
        invalid_count = len(preview_data) - valid_count

        # Mostrar notificación y mantener wizard abierto para mostrar botón "Confirmar"
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Análisis Completado'),
                'message': _('Productos válidos: %d, Con errores: %d') % (valid_count, invalid_count),
                'type': 'success' if invalid_count == 0 else 'warning',
                'sticky': False,  # No mantener visible para permitir click en Confirmar
                'fadeout': 'quick',
                'forceReload': True,  # Forzar recarga para actualizar botones
            },
        }

    def action_confirm_import(self):
        """Confirm import and create products - OPTIMIZED with batch processing"""
        self.ensure_one()

        if not self.preview_ids:
            raise UserError(_("No hay productos para importar"))

        valid_products = self.preview_ids.filtered(lambda p: p.is_valid)
        if not valid_products:
            raise UserError(_("No hay productos válidos para importar. Corrija los errores primero."))

        # PRECARGAR CATEGORÍAS - Una sola consulta con FUZZY MATCH
        unique_categ_names = set(valid_products.mapped('categ_name'))
        categories_cache = {}
        categories_created = []
        categories_matched = []
        
        for categ_name in unique_categ_names:
            if categ_name:
                # 1. Intentar fuzzy match con categorías existentes
                category = find_best_match_category(categ_name, self.env['product.category'])
                
                if category:
                    # Encontró categoría similar
                    categories_matched.append((categ_name, category.name))
                else:
                    # No encontró similar, crear nueva
                    category = self.env['product.category'].create({'name': categ_name})
                    categories_created.append(categ_name)
                
                categories_cache[categ_name] = category

        # PRECARGAR CATEGORÍAS PdV con FUZZY MATCH
        unique_pos_categ_names = set(valid_products.mapped('pos_categ_name'))
        pos_categories_cache = {}
        pos_categories_created = []
        pos_categories_matched = []
        
        for pos_categ_name in unique_pos_categ_names:
            if pos_categ_name and 'pos.category' in self.env:
                # 1. Intentar fuzzy match con categorías existentes
                pos_category = find_best_match_category(pos_categ_name, self.env['pos.category'])
                
                if pos_category:
                    # Encontró categoría similar
                    pos_categories_matched.append((pos_categ_name, pos_category.name))
                else:
                    # No encontró similar, crear nueva
                    pos_category = self.env['pos.category'].create({'name': pos_categ_name})
                    pos_categories_created.append(pos_categ_name)
                
                pos_categories_cache[pos_categ_name] = pos_category

        # CREACIÓN MASIVA DE PRODUCTOS - Batch processing
        product_vals_list = []
        products_to_quant = []
        
        for preview in valid_products:
            categ_id = self.env.ref('product.product_category_all').id
            if preview.categ_name and preview.categ_name in categories_cache:
                categ_id = categories_cache[preview.categ_name].id

            pos_categ_id = False
            if preview.pos_categ_name and preview.pos_categ_name in pos_categories_cache:
                pos_categ_id = pos_categories_cache[preview.pos_categ_name].id

            product_vals = {
                'name': preview.name,
                'default_code': preview.default_code,
                'barcode': preview.barcode or False,
                'list_price': preview.list_price,
                'standard_price': preview.standard_price,
                'type': preview.product_type,
                'categ_id': categ_id,
                'tracking': preview.tracking,
                'available_in_pos': preview.available_in_pos,
            }

            if preview.pos_description:
                product_vals['description_sale'] = preview.pos_description

            if pos_categ_id:
                product_vals['pos_categ_id'] = pos_categ_id

            product_vals_list.append(product_vals)
            
            # Guardar referencia para inventario
            if preview.qty_on_hand > 0:
                products_to_quant.append((len(product_vals_list) - 1, preview.qty_on_hand))

        # Creación masiva en una sola operación
        created_products = self.env['product.product'].create(product_vals_list)

        # APLICAR INVENTARIO MASIVO - Batch processing
        if products_to_quant and self.location_id:
            quant_vals_list = []
            for idx, qty in products_to_quant:
                product = created_products[idx]
                quant_vals_list.append({
                    'product_id': product.id,
                    'location_id': self.location_id.id,
                    'inventory_quantity': qty,
                })

            # Creación masiva de quants
            quants = self.env['stock.quant'].with_context(inventory_mode=True).create(quant_vals_list)
            # Aplicar inventario
            for quant in quants:
                quant.action_apply_inventory()

        self.write({'state': 'done'})

        # Construir mensaje detallado
        message = _('Se crearon %d productos exitosamente.') % len(created_products)
        
        if categories_matched:
            matched_list = ', '.join([f'"{orig}" → "{match}"' for orig, match in categories_matched])
            message += f'\n\n📁 Categorías reutilizadas (fuzzy match): {matched_list}'
        
        if categories_created:
            created_list = ', '.join(categories_created)
            message += f'\n📁 Categorías creadas: {created_list}'

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Proceso Completado'),
                'message': message,
                'type': 'success',
                'next': {'type': 'ir.actions.act_window_close'},
                'sticky': True,  # Mantener notificación visible para que el usuario vea el detalle
            }
        }


class ProductMassImportPreview(models.TransientModel):
    _name = 'product.mass.import.preview'
    _description = 'Vista Previa de Importación de Productos'
    _order = 'row_number'

    wizard_id = fields.Many2one('product.mass.import.wizard', string='Wizard', ondelete='cascade')
    row_number = fields.Integer(string='Fila')
    default_code = fields.Char(string='Referencia Interna')
    name = fields.Char(string='Nombre del Producto')
    pos_description = fields.Char(string='Descripción para PdV')
    barcode = fields.Char(string='Código de Barras')
    available_in_pos = fields.Boolean(string='Disponible en PdV')
    categ_name = fields.Char(string='Categoría de Producto')
    pos_categ_name = fields.Char(string='Categoría de PdV')
    list_price = fields.Float(string='Precio de Venta')
    standard_price = fields.Float(string='Precio de Costo')
    qty_on_hand = fields.Float(string='Cantidad a la Mano')
    product_type = fields.Selection([
        ('consu', 'Bienes (Almacenable/Consumible)'),
        ('service', 'Servicio'),
        ('combo', 'Combo'),
    ], string='Tipo de Producto', default='consu')
    tracking = fields.Selection([
        ('none', 'Ninguno'),
        ('lot', 'Por Lote'),
        ('serial', 'Por Número de Serie'),
    ], string='Trazabilidad', default='none')
    error_message = fields.Text(string='Errores')
    is_valid = fields.Boolean(string='Válido')
