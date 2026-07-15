# -*- coding: utf-8 -*-

import base64
import io
import logging
from datetime import datetime, date

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
except ImportError:
    openpyxl = None
    _logger.warning("No se encontró el módulo 'openpyxl'. Instálalo para usar la importación Excel.")


class PurchaseImportExcelWizard(models.TransientModel):
    _name = 'purchase.import.excel.wizard'
    _description = 'Importar Compras desde Excel'

    # ─── Modos de operación ───────────────────────────────────────────────────
    import_mode = fields.Selection([
        ('new', 'Crear nueva Orden de Compra'),
        ('existing', 'Añadir líneas a OC existente'),
    ], string='Modo de importación', default='new', required=True)

    purchase_order_id = fields.Many2one(
        'purchase.order',
        string='Orden de Compra',
        help='Orden de compra a la que se agregarán las líneas (solo para modo "existente").',
    )

    # ─── Archivo ──────────────────────────────────────────────────────────────
    excel_file = fields.Binary(string='Archivo Excel (.xlsx)')
    excel_filename = fields.Char(string='Nombre del archivo')

    # ─── Preview / resultado ─────────────────────────────────────────────────
    preview_line_ids = fields.One2many(
        'purchase.import.excel.line',
        'wizard_id',
        string='Vista previa de líneas',
    )
    import_errors = fields.Text(string='Errores de importación', readonly=True)
    state = fields.Selection([
        ('upload', 'Cargar archivo'),
        ('preview', 'Revisar datos'),
        ('done', 'Importación completada'),
    ], default='upload', string='Estado')

    result_order_ids = fields.Many2many(
        'purchase.order',
        string='Órdenes generadas',
    )
    result_summary = fields.Text(string='Resumen', readonly=True)

    # ─── Helper: leer Excel ───────────────────────────────────────────────────
    def _read_excel(self, file_data):
        """Lee el archivo Excel y devuelve lista de dicts con los datos de cada fila."""
        if not openpyxl:
            raise UserError(_(
                "La librería 'openpyxl' no está instalada en el servidor Odoo.\n"
                "Ejecuta: pip install openpyxl"
            ))

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(base64.b64decode(file_data)),
                read_only=True, data_only=True
            )
        except Exception as e:
            raise UserError(_("No se pudo leer el archivo Excel: %s") % str(e))

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise UserError(_("El archivo Excel está vacío."))

        # Detectar cabecera (primera fila con contenido)
        header_row = None
        data_start = 0
        for i, row in enumerate(rows):
            non_empty = [c for c in row if c is not None and str(c).strip() != '']
            if len(non_empty) >= 3:
                header_row = [str(c).strip().lower() if c else '' for c in row]
                data_start = i + 1
                break

        if header_row is None:
            raise UserError(_("No se encontró una fila de encabezado válida."))

        # Mapeo de columnas (flexible, acepta variantes en español/inglés)
        column_map = self._detect_columns(header_row)
        _logger.info("Mapa de columnas detectado: %s", column_map)

        records = []
        for row_idx, row in enumerate(rows[data_start:], start=data_start + 2):
            if all(c is None or str(c).strip() == '' for c in row):
                continue  # saltar filas vacías

            rec = {}
            for field_key, col_idx in column_map.items():
                if col_idx is not None and col_idx < len(row):
                    val = row[col_idx]
                    rec[field_key] = val if val is not None else ''
                else:
                    rec[field_key] = ''
            rec['_row'] = row_idx
            records.append(rec)

        wb.close()
        return records

    def _detect_columns(self, header):
        """
        Detecta las columnas del Excel a partir del encabezado.
        Devuelve un dict {field_key: col_index}.
        """
        synonyms = {
            'supplier': [
                'proveedor', 'supplier', 'vendor', 'razón social', 'razon social',
                'nombre proveedor', 'partner',
            ],
            'date_order': [
                'fecha', 'fecha compra', 'fecha de compra', 'date', 'date order',
                'fecha orden', 'fecha pedido',
            ],
            'product': [
                'producto', 'product', 'artículo', 'articulo', 'item',
                'descripción', 'descripcion', 'description', 'nombre producto',
                'product name',
            ],
            'product_code': [
                'código', 'codigo', 'code', 'ref', 'referencia', 'sku',
                'internal reference', 'referencia interna', 'product code',
            ],
            'qty': [
                'cantidad', 'qty', 'quantity', 'cant', 'unidades', 'units',
            ],
            'price_unit': [
                'costo', 'precio', 'price', 'cost', 'precio unitario',
                'unit price', 'precio costo', 'costo unitario',
            ],
            'tax': [
                'impuesto', 'tax', 'iva', 'taxes', 'alícuota', 'alicuota',
                'tax %', 'impuesto %',
            ],
            'margin': [
                'margen', 'margin', 'margen %', 'margin %',
                'porcentaje margen', 'markup', 'mark up',
            ],
            'lot_number': [
                'lote', 'lot', 'número de lote', 'numero de lote', 'lot number',
                'lot/serie', 'lot/serial', 'lote/serie', 'serie', 'serial',
                'nro lote', 'n° lote',
            ],
            'expiry_date': [
                'vencimiento', 'caducidad', 'fecha vencimiento', 'fecha caducidad',
                'expiry', 'expiry date', 'expiration', 'expiration date',
                'best before', 'fecha de vencimiento', 'fecha de caducidad',
                'use by', 'use by date',
            ],
            'uom': [
                'unidad', 'uom', 'unit', 'unit of measure', 'unidad de medida',
                'um', 'u.m.',
            ],
            'currency': [
                'moneda', 'currency', 'divisa',
            ],
        }

        col_map = {k: None for k in synonyms}
        for col_idx, header_val in enumerate(header):
            hv = header_val.lower().strip()
            for field_key, names in synonyms.items():
                if col_map[field_key] is None and hv in names:
                    col_map[field_key] = col_idx
                    break
                # coincidencia parcial como fallback
                if col_map[field_key] is None:
                    for name in names:
                        if name in hv or hv in name:
                            col_map[field_key] = col_idx
                            break
        return col_map

    # ─── Paso 1: Cargar y previsualizar ───────────────────────────────────────
    def action_preview(self):
        """Lee el Excel, valida datos y muestra la vista previa."""
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_("Por favor cargue un archivo Excel."))

        records = self._read_excel(self.excel_file)
        if not records:
            raise UserError(_("No se encontraron datos en el archivo."))

        # Limpiar líneas previas
        self.preview_line_ids.unlink()

        errors = []
        lines_vals = []

        for rec in records:
            row_num = rec.get('_row', '?')
            vals, row_errors = self._parse_record(rec, row_num)
            if row_errors:
                errors.extend(row_errors)
            if vals:
                vals['wizard_id'] = self.id
                lines_vals.append(vals)

        if lines_vals:
            self.env['purchase.import.excel.line'].create(lines_vals)

        self.import_errors = '\n'.join(errors) if errors else False
        self.state = 'preview'

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _parse_record(self, rec, row_num):
        """Parsea un registro del Excel y devuelve (vals_dict, [errores])."""
        errors = []
        vals = {}

        # ── Proveedor ──
        supplier_name = str(rec.get('supplier', '') or '').strip()
        if supplier_name:
            partner = self.env['res.partner'].search([
                ('name', 'ilike', supplier_name),
                ('supplier_rank', '>', 0),
            ], limit=1)
            if not partner:
                # Buscar también sin filtro de proveedor
                partner = self.env['res.partner'].search([
                    ('name', 'ilike', supplier_name),
                ], limit=1)
            if partner:
                vals['partner_id'] = partner.id
                vals['partner_name'] = partner.name
            else:
                errors.append(
                    _("Fila %s: Proveedor '%s' no encontrado en el sistema.") % (row_num, supplier_name)
                )
                vals['partner_name'] = supplier_name

        # ── Fecha ──
        date_val = rec.get('date_order', '')
        if date_val:
            parsed_date = self._parse_date(date_val)
            if parsed_date:
                vals['date_order'] = parsed_date
            else:
                errors.append(
                    _("Fila %s: Fecha '%s' no reconocida (use DD/MM/AAAA).") % (row_num, date_val)
                )
        else:
            vals['date_order'] = fields.Date.today()

        # ── Producto ──
        product_code = str(rec.get('product_code', '') or '').strip()
        product_name = str(rec.get('product', '') or '').strip()

        product = None
        if product_code:
            product = self.env['product.product'].search([
                '|',
                ('default_code', '=', product_code),
                ('default_code', 'ilike', product_code),
            ], limit=1)
        if not product and product_name:
            product = self.env['product.product'].search([
                '|',
                ('name', '=', product_name),
                ('name', 'ilike', product_name),
            ], limit=1)

        if product:
            vals['product_id'] = product.id
            vals['product_name'] = product.display_name
        else:
            search_term = product_code or product_name
            if search_term:
                errors.append(
                    _("Fila %s: Producto '%s' no encontrado. Se dejará en blanco.") % (row_num, search_term)
                )
            vals['product_name'] = product_name or product_code

        # ── Cantidad ──
        qty_val = rec.get('qty', 1)
        try:
            vals['product_qty'] = float(str(qty_val).replace(',', '.') or '1')
        except (ValueError, TypeError):
            vals['product_qty'] = 1.0
            errors.append(_("Fila %s: Cantidad '%s' inválida, se usará 1.") % (row_num, qty_val))

        # ── Precio / Costo ──
        price_val = rec.get('price_unit', 0)
        try:
            vals['price_unit'] = float(str(price_val).replace(',', '.') or '0')
        except (ValueError, TypeError):
            vals['price_unit'] = 0.0
            errors.append(_("Fila %s: Precio '%s' inválido.") % (row_num, price_val))

        # ── Impuesto ──
        tax_val = rec.get('tax', '')
        tax_str = str(tax_val).strip().replace(',', '.').replace('%', '').strip() if tax_val else ''
        if tax_str:
            try:
                tax_pct = float(tax_str)
                # Buscar impuesto por nombre
                tax = self.env['account.tax'].search([
                    ('type_tax_use', 'in', ['purchase', 'all']),
                    ('amount', '=', tax_pct),
                    ('company_id', '=', self.env.company.id),
                ], limit=1)
                if tax:
                    vals['tax_id'] = tax.id
                    vals['tax_name'] = tax.name
                else:
                    vals['tax_name'] = '%s%%' % tax_pct
                vals['tax_amount'] = tax_pct
            except (ValueError, TypeError):
                # Buscar por nombre
                tax = self.env['account.tax'].search([
                    ('name', 'ilike', tax_str),
                    ('type_tax_use', 'in', ['purchase', 'all']),
                    ('company_id', '=', self.env.company.id),
                ], limit=1)
                if tax:
                    vals['tax_id'] = tax.id
                    vals['tax_name'] = tax.name
                    vals['tax_amount'] = tax.amount
                else:
                    vals['tax_name'] = tax_str

        # ── Margen ──
        margin_val = rec.get('margin', '')
        margin_str = str(margin_val).strip().replace(',', '.').replace('%', '').strip() if margin_val else ''
        if margin_str:
            try:
                vals['margin_percent'] = float(margin_str)
            except (ValueError, TypeError):
                vals['margin_percent'] = 0.0

        # ── Número de lote ──
        lot_val = rec.get('lot_number', '')
        vals['lot_number'] = str(lot_val).strip() if lot_val else ''

        # ── Fecha de caducidad ──
        exp_val = rec.get('expiry_date', '')
        if exp_val:
            parsed_exp = self._parse_date(exp_val)
            if parsed_exp:
                vals['expiry_date'] = parsed_exp
            else:
                errors.append(
                    _("Fila %s: Fecha de caducidad '%s' no reconocida.") % (row_num, exp_val)
                )

        # ── UoM ──
        uom_name = str(rec.get('uom', '') or '').strip()
        if uom_name:
            uom = self.env['uom.uom'].search([('name', 'ilike', uom_name)], limit=1)
            if uom:
                vals['product_uom_id'] = uom.id

        # Si no hay producto ni proveedor informado, skip silencioso
        if not product_name and not product_code and not supplier_name:
            return None, []

        vals['import_status'] = 'error' if errors else 'ok'
        vals['import_notes'] = '\n'.join(errors) if errors else ''

        return vals, errors

    def _parse_date(self, value):
        """Intenta parsear distintos formatos de fecha y devuelve string 'YYYY-MM-DD' o None."""
        if not value:
            return None

        if isinstance(value, (datetime, date)):
            if isinstance(value, datetime):
                return value.date().strftime('%Y-%m-%d')
            return value.strftime('%Y-%m-%d')

        value_str = str(value).strip()
        formats = [
            '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d',
            '%d/%m/%y', '%d-%m-%y', '%m/%d/%Y', '%m-%d-%Y',
            '%d.%m.%Y', '%d.%m.%y',
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value_str, fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
        return None

    # ─── Paso 2: Importar ─────────────────────────────────────────────────────
    def action_import(self):
        """Crea las órdenes de compra a partir de las líneas validadas."""
        self.ensure_one()

        valid_lines = self.preview_line_ids.filtered(
            lambda l: l.product_id and l.import_status != 'skip'
        )
        if not valid_lines:
            raise UserError(_(
                "No hay líneas válidas para importar. "
                "Asegúrese de que los productos existan en el sistema."
            ))

        orders_created = self.env['purchase.order']

        if self.import_mode == 'existing' and self.purchase_order_id:
            # ── Agregar a orden existente ──
            self._add_lines_to_order(self.purchase_order_id, valid_lines)
            orders_created |= self.purchase_order_id
        else:
            # ── Crear orden(es) nueva(s) agrupadas por proveedor + fecha ──
            groups = {}
            for line in valid_lines:
                key = (line.partner_id.id if line.partner_id else 0,
                       str(line.date_order or fields.Date.today()))
                groups.setdefault(key, []).append(line)

            for (partner_id, date_order), group_lines in groups.items():
                if not partner_id:
                    continue  # sin proveedor no se puede crear OC
                order = self._create_purchase_order(partner_id, date_order, group_lines)
                orders_created |= order

        if not orders_created:
            raise UserError(_(
                "No se pudieron crear órdenes. Verifique que todos los proveedores existan en el sistema."
            ))

        self.result_order_ids = orders_created
        self.result_summary = _(
            "✅ Importación completada.\n"
            "%d orden(es) procesada(s).\n"
            "%d línea(s) importada(s)."
        ) % (len(orders_created), len(valid_lines))
        self.state = 'done'

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _create_purchase_order(self, partner_id, date_order, lines):
        """Crea una orden de compra y sus líneas."""
        partner = self.env['res.partner'].browse(partner_id)
        order_vals = {
            'partner_id': partner_id,
            'date_order': date_order,
            'state': 'draft',
            'company_id': self.env.company.id,
            'notes': _('Importado desde Excel - %s') % fields.Datetime.now(),
        }
        order = self.env['purchase.order'].create(order_vals)
        self._add_lines_to_order(order, lines)
        return order

    def _add_lines_to_order(self, order, lines):
        """Agrega líneas a una orden de compra existente."""
        for line in lines:
            if not line.product_id:
                continue

            product = line.product_id
            # UoM
            uom = line.product_uom_id or product.uom_po_id or product.uom_id

            # Impuesto
            taxes = self.env['account.tax']
            if line.tax_id:
                taxes = line.tax_id
            else:
                taxes = product.supplier_taxes_id.filtered(
                    lambda t: t.company_id == self.env.company
                )

            line_vals = {
                'order_id': order.id,
                'product_id': product.id,
                'name': product.name,
                'product_qty': line.product_qty,
                'price_unit': line.price_unit,
                'product_uom': uom.id,
                'taxes_id': [(6, 0, taxes.ids)],
                'date_planned': line.date_order or fields.Date.today(),
            }
            po_line = self.env['purchase.order.line'].create(line_vals)

            # Actualizar precio de venta con margen
            if line.margin_percent and line.margin_percent > 0:
                sale_price = line.price_unit * (1 + line.margin_percent / 100.0)
                product.write({'list_price': sale_price})

            # Guardar datos de lote en la línea de importación (para usarlos en recepción)
            if line.lot_number or line.expiry_date:
                line.write({'purchase_line_id': po_line.id})

    # ─── Paso 3: Abrir órdenes ────────────────────────────────────────────────
    def action_view_orders(self):
        """Abre las órdenes de compra creadas."""
        self.ensure_one()
        if not self.result_order_ids:
            return {'type': 'ir.actions.act_window_close'}

        if len(self.result_order_ids) == 1:
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'purchase.order',
                'res_id': self.result_order_ids.id,
                'view_mode': 'form',
                'target': 'current',
            }
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'purchase.order',
            'domain': [('id', 'in', self.result_order_ids.ids)],
            'view_mode': 'list,form',
            'target': 'current',
        }

    def action_close(self):
        return {'type': 'ir.actions.act_window_close'}

    # ─── Plantilla Excel ──────────────────────────────────────────────────────
    def action_download_template(self):
        """Genera y descarga una plantilla Excel de ejemplo."""
        if not openpyxl:
            raise UserError(_("La librería 'openpyxl' no está instalada."))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Importar Compras'

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Colores
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        required_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        optional_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        lot_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, name='Calibri', size=10)
        cell_font = Font(name='Calibri', size=10)
        thin = Side(border_style='thin', color='BFBFBF')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center')

        headers = [
            ('Proveedor *', 25, 'required'),
            ('Fecha Compra *', 15, 'required'),
            ('Código Producto', 18, 'optional'),
            ('Producto *', 30, 'required'),
            ('Cantidad *', 12, 'required'),
            ('Costo Unitario *', 16, 'required'),
            ('Impuesto %', 12, 'optional'),
            ('Margen %', 12, 'optional'),
            ('N° Lote / Serie', 20, 'lot'),
            ('Fecha Caducidad', 16, 'lot'),
            ('Unidad de Medida', 18, 'optional'),
        ]

        # Fila de leyenda
        ws.merge_cells('A1:K1')
        legend_cell = ws['A1']
        legend_cell.value = (
            '📋 PLANTILLA IMPORTACIÓN DE COMPRAS | '
            'Columnas marcadas * son obligatorias | '
            '🟡 Lote/Caducidad: completar solo si el producto usa lotes/series'
        )
        legend_cell.font = Font(name='Calibri', size=9, italic=True, color='444444')
        legend_cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        legend_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 30

        # Cabecera
        for col_num, (header_text, col_width, col_type) in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_num, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(col_num)].width = col_width
        ws.row_dimensions[2].height = 22

        # Datos de ejemplo
        example_rows = [
            ['Proveedor Ejemplo S.A.', '15/01/2025', 'PROD-001', 'Ibuprofeno 400mg x 20 comp',
             100, 250.50, 21, 30, 'LOT2025A', '31/12/2025', 'Caja'],
            ['Proveedor Ejemplo S.A.', '15/01/2025', 'PROD-002', 'Amoxicilina 500mg x 12 caps',
             50, 480.00, 21, 25, '', '', 'Caja'],
            ['Otro Proveedor S.R.L.', '20/01/2025', '', 'Producto Sin Lote',
             200, 15.75, 10.5, 20, '', '', 'Unidad'],
        ]

        fill_map = {'required': required_fill, 'optional': optional_fill, 'lot': lot_fill}

        for row_num, row_data in enumerate(example_rows, start=3):
            for col_num, value in enumerate(row_data, start=1):
                col_type = headers[col_num - 1][2]
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = cell_font
                cell.fill = fill_map.get(col_type, optional_fill)
                cell.border = border
                cell.alignment = Alignment(vertical='center')
            ws.row_dimensions[row_num].height = 18

        # Instrucciones en segunda hoja
        ws2 = wb.create_sheet('Instrucciones')
        instructions = [
            ('INSTRUCCIONES DE USO', True),
            ('', False),
            ('CAMPOS OBLIGATORIOS:', True),
            ('• Proveedor: Nombre exacto o parcial del proveedor (debe existir en Odoo)', False),
            ('• Fecha Compra: Formato DD/MM/AAAA (ej: 15/01/2025)', False),
            ('• Producto: Nombre o código de referencia del producto en Odoo', False),
            ('• Cantidad: Número entero o decimal (usar punto como separador decimal)', False),
            ('• Costo Unitario: Precio de compra sin impuestos', False),
            ('', False),
            ('CAMPOS OPCIONALES:', True),
            ('• Código Producto: Referencia interna del producto (más preciso que el nombre)', False),
            ('• Impuesto %: Porcentaje de IVA/impuesto (ej: 21 para 21%). Debe existir en Odoo', False),
            ('• Margen %: Porcentaje de ganancia sobre el costo para calcular precio de venta', False),
            ('  Ejemplo: Si costo=100 y margen=30, el precio de venta se seteará en 130', False),
            ('• Unidad de Medida: Unidad de compra del producto', False),
            ('', False),
            ('LOTE Y CADUCIDAD (solo para productos que usen lotes/series):', True),
            ('• N° Lote / Serie: Número de lote o número de serie del producto', False),
            ('• Fecha Caducidad: Fecha de vencimiento del lote (DD/MM/AAAA)', False),
            ('  Estos datos se asignan automáticamente al recepcionar la orden en Inventario', False),
            ('', False),
            ('NOTAS IMPORTANTES:', True),
            ('• Se pueden importar múltiples proveedores en el mismo archivo', False),
            ('• Se creará una orden de compra por cada combinación proveedor+fecha única', False),
            ('• Los productos deben existir en Odoo antes de importar', False),
            ('• La primera fila con datos es el encabezado; las columnas se detectan automaticamente', False),
        ]

        ws2.column_dimensions['A'].width = 80
        for i, (text, bold) in enumerate(instructions, start=1):
            cell = ws2.cell(row=i, column=1, value=text)
            cell.font = Font(name='Calibri', size=10, bold=bold,
                             color='1F4E79' if bold else '000000')
            ws2.row_dimensions[i].height = 16

        # Exportar a bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_data = base64.b64encode(buffer.read())

        # Crear adjunto temporal
        attachment = self.env['ir.attachment'].create({
            'name': 'plantilla_importar_compras.xlsx',
            'type': 'binary',
            'datas': file_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%d?download=true' % attachment.id,
            'target': 'new',
        }


class PurchaseImportExcelLine(models.TransientModel):
    _name = 'purchase.import.excel.line'
    _description = 'Línea de importación Excel'

    wizard_id = fields.Many2one(
        'purchase.import.excel.wizard',
        string='Wizard',
        required=True,
        ondelete='cascade',
    )

    # ── Datos básicos ──
    partner_id = fields.Many2one('res.partner', string='Proveedor')
    partner_name = fields.Char(string='Proveedor (texto)')
    date_order = fields.Date(string='Fecha Compra')

    # ── Producto ──
    product_id = fields.Many2one('product.product', string='Producto')
    product_name = fields.Char(string='Producto (texto)')
    product_qty = fields.Float(string='Cantidad', default=1.0)
    product_uom_id = fields.Many2one('uom.uom', string='Unidad')
    price_unit = fields.Float(string='Costo Unitario')

    # ── Impuesto y margen ──
    tax_id = fields.Many2one('account.tax', string='Impuesto')
    tax_name = fields.Char(string='Impuesto (texto)')
    tax_amount = fields.Float(string='Impuesto %')
    margin_percent = fields.Float(string='Margen %')
    sale_price_computed = fields.Float(
        string='Precio Venta',
        compute='_compute_sale_price',
        store=False,
    )

    # ── Lote y caducidad ──
    lot_number = fields.Char(string='N° Lote / Serie')
    expiry_date = fields.Date(string='Fecha Caducidad')

    # ── Control ──
    import_status = fields.Selection([
        ('ok', '✅ OK'),
        ('warning', '⚠️ Advertencia'),
        ('error', '❌ Error'),
        ('skip', '⏭️ Omitir'),
    ], string='Estado', default='ok')
    import_notes = fields.Text(string='Notas')

    # ── Referencia a la línea creada ──
    purchase_line_id = fields.Many2one(
        'purchase.order.line',
        string='Línea OC creada',
    )

    @api.depends('price_unit', 'margin_percent')
    def _compute_sale_price(self):
        for rec in self:
            if rec.margin_percent:
                rec.sale_price_computed = rec.price_unit * (1 + rec.margin_percent / 100.0)
            else:
                rec.sale_price_computed = rec.price_unit

    def action_skip(self):
        self.write({'import_status': 'skip'})

    def action_include(self):
        self.write({'import_status': 'ok'})
